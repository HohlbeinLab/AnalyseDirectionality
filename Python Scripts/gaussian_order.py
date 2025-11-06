import os
import pickle
import sys
from typing import Any, Iterable

import matplotlib as mpl
import numpy as np
import matplotlib.pyplot as plt
import scipy
from scipy.optimize import curve_fit
from scipy.signal import find_peaks
from tqdm import tqdm
import warnings
from openpyxl import Workbook
from astropy.stats import circmean, circstd
from multiprocessing import Pool
import time
import argparse


class FittingClass:
    """
    Class structure has been utilised to support multithreading more efficiently
    Without this implementation, a naive approach would mean heavy slowdown for small samples
    The class takes in a set of constants that are used to define the x-axis angles
    as well as certain fitting parameters.
    """

    def __init__(self, angles, p_per_a, en, st, pts, prominence, min_width, min_distance, max_gausses=4, plot=False):
        self.min_value = None
        self.arr = None
        self.row_idx = None
        self.plot = plot
        self.max_gausses = max_gausses
        self.angles = angles
        self.p_per_a = p_per_a
        self.en = en
        self.st = st
        self.pts = pts
        self.prominence = prominence
        self.min_width = min_width
        self.min_distance = min_distance

    def fit_gaussian(self) -> tuple:
        ax_plot = gauss_fig = ax_residual = None

        if self.plot:
            gauss_fig = plt.figure(1)
            ax_plot = gauss_fig.add_axes((.1, .3, .8, .6))
            ax_residual = gauss_fig.add_axes((.1, .1, .8, .2))

        gauss_offset = np.argmin(self.arr)
        rolled_arr = np.roll(self.arr, -gauss_offset)
        params = covariance = residuals = False
        peaks, details = find_peaks(rolled_arr,
                                    prominence=[self.prominence*np.max(rolled_arr)],
                                    width=[self.min_width],
                                    rel_height=0.33,
                                    distance=self.min_distance)

        gausses = min(self.max_gausses + 1, len(peaks) + 1)
        r = np.arange(0, gausses, 1)
        if peaks.shape[0] > gausses:
            r = details["widths"].argsort()[::-1][:gausses]
        elif peaks.shape[0] == 0: # no peaks detected
            return [[np.nan, np.nan, np.nan]], [[np.nan, np.nan, np.nan]], np.nan, np.nan

        # mean, amplitude, width
        guess = []
        bounds = [[], []]

        for i in r:
            if i < peaks.shape[0]:

                guess += [peaks[i] / self.p_per_a, rolled_arr[peaks[i]], details["widths"][i]/2]
                bounds[0] += [peaks[i] / self.p_per_a * 0.95, 0.05 * rolled_arr[peaks[i]], details["widths"][i] * 0.05]
                bounds[1] += [peaks[i] / self.p_per_a * 1.05+0.001, 1.5 * rolled_arr[peaks[i]]+0.0001, details["widths"][i] * 5+0.001]
            else:
                guess += [self.angles[-1] // 2, np.max(rolled_arr) * 0.25, self.angles[-1] / 20]
                bounds[0] += [0, 0, 0]
                bounds[1] += [self.angles[-1], 1.1 * np.max(self.arr), self.angles[-1] / 4]

        try:
            params, covariance = curve_fit(self.fit_gauss, self.angles, rolled_arr, p0=guess, bounds=bounds)
            for i in range(0, len(params), 3):
                params[i] = (params[i] + gauss_offset / self.p_per_a) % 180

            shifted_params1 = params.copy()
            shifted_params2 = params.copy()
            if self.plot:

                for i in range(0, len(params), 3):
                    shifted_params1[i] -= self.en
                    shifted_params2[i] += self.en

                    ax_plot.plot(self.angles, gauss(self.angles, *shifted_params1[i:i + 3]), c='lightgreen')
                    ax_plot.plot(self.angles, gauss(self.angles, *shifted_params2[i:i + 3]), c='darkgreen')
                    #ax_plot.plot(angles, gauss(angles, *params[i:i + 3]), 'r-')

            sum_fit = (gauss(self.angles, *params) +
                       gauss(self.angles, *shifted_params1) +
                       gauss(self.angles, *shifted_params2))
            residuals = sum_fit - self.arr

            if self.plot:
                #pass
                ax_residual.plot(self.angles, residuals)
                for i in range(0, len(params), 3):
                    ax_plot.plot(self.angles, gauss(self.angles, *params[i:i + 3]), 'r-')

                #for i in range(0, len(guess), 3):
                #    guess[i] = (guess[i] + gauss_offset / self.p_per_a) % 180
                #    ax_plot.plot(angles, gauss(angles, *guess[i:i + 3]), 'g-')
                ax_plot.plot(self.angles, sum_fit, 'b-', label='sum fit')

        except RuntimeError as err:
            print(err)
        except Exception as err:
            print(self.arr)
            print(peaks)
            plt.figure()
            plt.title(f"{self.row_idx}")
            plt.plot(self.angles, self.arr, c='k', label="data")
            plt.plot(self.angles, rolled_arr, c='b', label="rolled")
            plt.legend()
            plt.show()
            raise err

        if self.plot:
            ax_plot.set_title(f"i: {self.row_idx}")
            ax_plot.stairs(self.arr, np.append(self.angles, self.angles[-1] * 2 - self.angles[-2]), color='k',
                           label="data")
            #ax_plot.plot(self.angles, rolled_arr, c='b', label="rolled")
            ax_residual.set_xlabel("Angle (°)")
            ax_plot.set_ylabel("Intensity (a.u)")
            ax_plot.scatter(self.angles[(peaks + gauss_offset) % len(self.angles)], rolled_arr[peaks], c="r",
                            label="peaks")
            #ax_plot.legend()

            ax_plot.set_xlim([self.st, self.en])
            ax_plot.set_xticklabels([])
            ax_residual.set_xlim(ax_plot.get_xlim())
            # ax_plot.set_ylim([0, 1])
            gauss_fig.savefig("testing.svg")
            gauss_fig.show()


        if type(params) is bool:
            return [[np.nan, np.nan, np.nan]], [[np.nan, np.nan, np.nan]], np.nan, np.nan
        else:
            params = [params[i: i + 3] for i in range(0, len(params), 3)]
            if self.plot:
                print(params)
            return *select(params, ((peaks + gauss_offset) / self.p_per_a) % 180, self.min_value), np.sqrt(
                np.diag(covariance)), np.sum(np.power(residuals, 2))

    def fit_gauss(self, x, *params):
        params = list(params)
        shifted_params1 = params[:]
        shifted_params2 = params[:]
        for i in range(0, len(params), 3):
            shifted_params1[i] -= self.en
            shifted_params2[i] += self.en
        return gauss(x, *params) + gauss(x, *shifted_params1) + gauss(x, *shifted_params2)

    def fit(self, args):
        d, index, idx_5 = args
        min_val = np.min(d)
        d -= min_val
        self.arr = d
        self.min_value = min_val
        self.row_idx = index
        return *self.fit_gaussian(), self.min_value, idx_5, self.row_idx




def heatmap(data_imshow, ax_imshow, cbarlabel="", cbar_yticks=None, **kwargs):
    # Plot the heatmap
    heatmap_im = ax_imshow.imshow(data_imshow, **kwargs)

    # Create colorbar
    heatmap_cbar = ax_imshow.figure.colorbar(heatmap_im, ax=ax_imshow)
    heatmap_cbar.ax.set_ylabel(cbarlabel, rotation=-90, va="bottom")
    if cbar_yticks:
        heatmap_cbar.ax.set_yticks(cbar_yticks)
    ax_imshow.tick_params(which="minor", bottom=False, left=False)
    ax_imshow.tick_params(which="major", bottom=False, left=False)
    ax_imshow.set_xticklabels([])
    ax_imshow.set_yticklabels([])

    return heatmap_im, heatmap_cbar


def calc_order_param(angs: np.ndarray, ints: np.ndarray, stds: np.ndarray, centre_value: list) -> list:
    angs = (angs + 90) % 180

    orders = []
    ord_stds = []
    ord_ints = []

    for angle, intens, std in centre_value:
        angle = (angle + 90) % 180
        cos_list_ang = np.power(np.cos(np.radians(angs - angle)), 2)
        cos_list_std = np.power(np.cos(np.radians(stds - std)), 2)
        orders.append(2 * (np.nanmean(cos_list_ang) - 0.5))
        ord_stds.append(2 * (np.nanmean(cos_list_std) - 0.5))
        ord_ints.append(np.nanstd(np.abs(ints - intens)))
    return [np.nanmean(orders), np.nanmean(ord_stds), np.nanmean(ord_ints)]


def flatten(lst: Iterable[Iterable]) -> Iterable:
    flat_list = []
    for r in lst:
        flat_list += r
    return flat_list


def calc_order_param_weighted(angs: Iterable, ints: Iterable, centre_values: Iterable):
    angs = (np.array(angs) + 90) % 180
    orders = []
    intensities = []
    for angle, intens, _ in centre_values:
        if np.isnan(angle):
            orders.append(0)
            intensities.append(0)
        else:
            angle = (angle + 90) % 180
            cos_list_ang = np.power(np.cos(np.radians(angs - angle)), 2)
            orders.append(2 * (weighted_nanmean(cos_list_ang, weights=np.power(ints, 2)) - 0.5))
            intensities.append(intens)
    return weighted_nanmean(orders, weights=np.power(intensities, 2))


def select(pars: list, peaks: np.ndarray, min_value: int) -> [tuple | list]:
    if np.isnan(pars[0][0]):
        return pars
    if any(peaks < 1):
        peaks = np.hstack((peaks, [180]))
    # Determine the signal and bg gausses
    # One min in each
    # mean, amplitude, width
    # angle, intensity, std

    if len(pars) == 1:
        return pars, [[]]
    else:
        sig = []
        bg = []

        for a, i, s in pars:
            if (all([si < i for sa, si, _ in pars if 0 < abs(a - sa) < 5])
                    and (s > 0.2 and (any(abs(peaks - a) < 10)
                                      or (i > min_value * 0.01 and s < 20)))):
                sig.append([a % 180, i, s])
            else:
                bg.append([a % 180, i, s])
        # print(peaks % 180)
        # print(details)
        # print(sig, bg)
        return sig, bg


def gauss(x, *params):
    y = np.zeros_like(x)
    for i in range(0, len(params), 3):
        ctr, amp, wid = params[i:i + 3]
        y += amp * np.exp(-((x - ctr) / (2 * wid)) ** 2)
    return y


def match(peaks: list, mainpeaks: list, match_ang: int = 15) -> list:
    closest = []
    for a, _, _ in peaks:
        close = np.array(
            [[i, abs(np.sin(np.radians(a - p))), p] for i, p in enumerate(mainpeaks) if
             abs(np.sin(np.radians(a - p))) < np.sin(np.radians(match_ang))])
        if close.shape[0] == 0:
            closest.append(np.nan)
        else:
            close = close[close[:, 1].argsort()]
            closest.append(close[0, 0])
    return closest


def find_first(haystack: np.ndarray, needle) -> int:
    for i, item in enumerate(haystack):
        if needle == item:
            return i
    else:
        return -1


def weighted_nanmean(A: [Iterable], weights: [Any] = 1):
    if isinstance(A, list) and not A:
        return np.nan
    A = np.array(A)
    weights = np.array(weights)
    if np.nansum((~np.isnan(A)) * weights) != 0:
        return np.nansum(A * weights) / np.nansum((~np.isnan(A)) * weights)
    else:
        return 0


def weighted_nanstd(A, weights=None):
    average = weighted_nanmean(A, weights=weights)
    # Fast and numerically precise:
    variance = weighted_nanmean((A - average) ** 2, weights=weights)
    return np.sqrt(variance)


def angular_weighted_nanmean(A, weights=None):
    if not A:
        return np.nan
    A = np.array(A)
    if weights is not None:
        weights = np.array(weights)[~np.isnan(A)]
        return np.degrees(circmean(np.radians(A[~np.isnan(A)] * 2), weights=weights) / 2) % 180
    else:
        return scipy.stats.circmean(A, high=180)


def angular_weighted_nanstd(A, weights=None):
    if not A:
        return np.nan
    A = np.array(A)

    if weights is None:
        return scipy.stats.circstd(A, high=180)
    else:
        weights = np.array(weights)[~np.isnan(A)]
        return np.degrees(circstd(np.radians(A[~np.isnan(A)] * 2), weights=weights) / 2) % 180


def make_dir(*paths):
    if not os.path.isdir(os.path.join(*paths)):
        os.makedirs(os.path.join(*paths))


def visualise_fit(visualise_instance: FittingClass, RFT, row: int):
    print(f"Visualising row {row}")
    visualise_instance.plot = True
    if len(RFT.shape) > 1:
        return visualise_instance.fit((RFT[row, 8:], row, RFT[row, 5]))
    else:
        return visualise_instance.fit((RFT[8:], row, RFT[5]))

def characteristic_size(mean_std: float, window_size: int):
    return window_size/(0.4+24.16/(1+np.power(mean_std/0.91, 1.54)))
    #return window_size/(0.91*(np.power((24.16/(ppw-0.4))-1,1/1.54)))

def prepare_parser(parser):
    parser.add_argument("-match_angle", help="Range of angles that get merged", type=float, default=35.0)
    parser.add_argument("-max_neighbourhood", help="Maximum width to calculate WOP for", type=int, default=7)
    parser.add_argument("-filter_edges", help="Amount of windows to remove from each edge", type=int, default=0)
    parser.add_argument("-prominence", help="Percentage of max value above minimum that a peak must have", type=float, default=0.08)
    parser.add_argument("-min_peak_width", help="Minimum width (°) a peak must have", type=float, default=1.0)
    parser.add_argument("-min_distance", help="Minimum distance (°) between peaks ", type=float, default=5.0)
    parser.add_argument("-singlethreaded", "-s", help="Disable Multithreading", action="store_true")
    parser.add_argument("-image_format", help="Image format to export matplotlib graphs in", type=str, default="svg")
    parser.add_argument("-testing", help="Will display detailed information on a single window", type=int, default=0)
    parser.add_argument("-seed", help="Seeded value to use for Numpy", type=int, default=23452987)
    parser.add_argument("-show_graph", help="Shows graphs", action="store_true")
    parser.add_argument("-all_angles", help="Calculate WOP for individual clustered angles", action="store_true")
    parser.add_argument("-no_recalculation", "-r", help="Store intermediate fitting results in a pickle file", action="store_true")

    parser.add_argument("-core_path", "-c", help="Folder to use as input. Use '-a' for absolute path, otherwise will search in '.\\input\\<core_path>'", type=str, default = "", nargs="+")
    parser.add_argument("-absolute", "-a", help="Use the core path absolutely", action="store_true")
    parser.add_argument("-filenames", '-f', help="Filename to use as input (Extension CSV) or 'all' for all filenames. Separate filenames by ','. Setting 'all' will recursively search all folders not named 'results' or 'pickles'.", nargs="+", type=str, default="all")

    parser.add_argument("-IDE", help="Add when running from IDE to use internal settings", action="store_true")


def run(argument_string = "", max_neighbourhood=None, filter_edges=None, prominence=None, min_peak_width=None, min_distance=None, singlethreaded=None,
        image_format=None, testing=None, seed=None, show_graph=None, all_angles=None, no_recalculation=None, core_path=None, filenames=None, absolute=None):
    #argument_string = "-c C:\\Users\\gobes001\\LocalSoftware\\AnalyseDirectionality\\test_path\\input -absolute -f window250_15cm_coronal_top_A_crop"

    parser = argparse.ArgumentParser(description="Analyses software for RFT", epilog="Find information at https://github.com/HohlbeinLab/AnalyseDirectionality")
    prepare_parser(parser)

    if not argument_string:
        args = parser.parse_args()
    else:
        args = parser.parse_args(argument_string.split())

    if args.filenames != "all":
        args.filenames = " ".join(args.filenames).replace("\"", "").split(",")
    args.core_path = " ".join(args.core_path).replace("\"", "")


    buff = 0
    st = 0 - buff
    en = 180 + buff

    if args.IDE: # change defaults if needed
        args.testing = 0
        args.show_graph = False
        args.no_recalculation = True
        #args.core_path = "400rectangle"
        args.core_path = r"D:\Data\2025_KatjaDeadstopPaper\results\input"
        args.filenames = "all"
        args.absolute=True
        #args.filenames = ["window100_coronal_top_A_crop", "window300_coronal_top_A_crop"]


    if max_neighbourhood is not None:
        args.max_neighbourhood = int(max_neighbourhood)
    if filter_edges is not None:
        args.filter_edges = bool(filter_edges)
    if prominence is not None:
        args.prominence = float(prominence)
    if min_peak_width is not None:
        args.min_peak_width = float(min_peak_width)
    if min_distance is not None:
        args.min_distance = float(min_distance)
    if singlethreaded is not None:
        args.singlethreaded = bool(singlethreaded)
    if image_format is not None:
        args.image_format = str(image_format)
    if testing is not None:
        args.testing = int(testing)
    if seed is not None:
        args.seed = int(seed)
    if show_graph is not None:
        args.show_graph = bool(show_graph)
    if all_angles is not None:
        args.all_angles = bool(all_angles)
    if no_recalculation is not None:
        args.no_recalculation = bool(no_recalculation)
    if core_path is not None:
        args.core_path = str(core_path)
    if filenames is not None:
        args.filenames = filenames
    if absolute is not None:
        args.absolute = bool(absolute)


    if args.absolute:
        if not 'input' in args.core_path.split(os.sep):
            load_path = os.path.join(args.core_path, "input")
        else:
            load_path = args.core_path
            drive, path = os.path.splitdrive(args.core_path)
            args.core_path = os.path.join(drive, os.sep, *path.split(os.sep)[1:path.split(os.sep).index("input")])
    else:
        load_path = os.path.join(".", "input", args.core_path)


    np.random.seed(args.seed)
    excel_book = Workbook()
    excel_book.remove(excel_book.active)
    overview_sheet = excel_book.create_sheet("Overview")
    overview_sheet.append(["Name"])
    filepaths = []

    if not args.singlethreaded:
        thread_pool = Pool()

    if args.filenames == "all":
        for path, subdirs, files in os.walk(load_path):
            for file in files:
                if file.endswith(".csv"):
                    filepaths.append(os.path.join(path, file))
        if not filepaths:
            raise Exception(f"No files found in the path {load_path}")
    elif args.filenames:
        errors = []

        for name in args.filenames:
            if not os.path.isfile(os.path.join(load_path, f"{name}.csv")):
                errors.append(os.path.join(load_path, f"{name}.csv"))
            else:
                filepaths.append(os.path.join(load_path, f"{name}.csv"))
        if errors:
            raise Exception(f"Could not find the following files: {errors}")
    else:
        raise Exception("No files to process. Define 'names'")

    for filepath in filepaths:
        filename = os.path.basename(filepath)[:-4]
        print(f"Processing {filename}")
        if "\\" in filename or "/" in filename:  # Some POSIX fun
            filename = filename.split("\\")[-1].split("/")[-1][:-4]
        folder_path = os.path.dirname(filepath)
        if args.absolute:
            results_path = os.path.join(args.core_path, "results", *folder_path.split(os.sep)[args.core_path.count(os.sep) + 2:], filename)
            pickle_path = os.path.join(args.core_path, "pickles", *folder_path.split(os.sep)[args.core_path.count(os.sep) + 2:])
        else:
            results_path = os.path.join(".", "results", *folder_path.split(os.sep)[2:], filename)
            pickle_path = os.path.join(".", "pickles", *folder_path.split(os.sep)[2:])

        excel_sheet = excel_book.create_sheet(filename)
        overview_sheet.append([filename])

        make_dir(results_path)
        make_dir(pickle_path)
        make_dir(results_path, "csv")
        make_dir(results_path, "raw")
        stats_file = open(os.path.join(results_path, "csv", f"{filename}_all_stats.csv"), "w")

        # 0 1   2       3       4           5           6       7           8:
        # x	y	width	height	Max Index	Mask Median	Angle	Relevance   data:
        RFT = np.loadtxt(filepath, delimiter=',', skiprows=1)
        if len(RFT.shape) > 1:
            all_data = np.sum(RFT[:, 8:], axis=0)
            window_size = RFT[0, 2]
        else:
            all_data = RFT[8:]
            window_size = RFT[2]
            if not args.testing:
                raise Exception("More than 1 row of data must be included")

        p_per_a = all_data.shape[0] / 180
        pts = int((en - st) * p_per_a)
        angles = np.linspace(st, en, pts)

        offset = np.argmin(all_data)
        rolled_all = np.roll(all_data, -offset)
        main_peaks, all_details = find_peaks(rolled_all, prominence=[0.10], distance=10)
        fitting_class_instance = FittingClass(angles=angles, p_per_a=p_per_a, en=en, st=st, pts=pts, prominence=args.prominence, min_width=args.min_peak_width, min_distance=args.min_distance)

        pickle_filename = os.path.join(pickle_path, f"{filename}.pickle")

        if os.path.isfile(pickle_filename) and args.no_recalculation and not testing:
            with open(pickle_filename, 'rb') as pickle_file:
                results = pickle.load(pickle_file)
                print(rf"Loaded {pickle_filename} with {len(results)} rows")
        else:
            if args.testing:
                visualise_fit(fitting_class_instance, RFT, args.testing)
                exit()

            start_time = time.time()
            if args.singlethreaded:
                results = []
                for i in tqdm(range(RFT.shape[0])):
                    results.append(fitting_class_instance.fit((RFT[i, 8:], i, RFT[i, 5])))
            else:
                results = list(
                    tqdm(thread_pool.imap_unordered(fitting_class_instance.fit, [(RFT[i, 8:], i, RFT[i, 5]) for i in
                                                                                range(RFT.shape[0])]), total=RFT.shape[0]))
            print(f"Time taken: {time.time() - start_time}")
            with open(pickle_filename, 'wb') as pickle_file:
                print(f"Saving pickle file to {pickle_filename}")
                pickle.dump(results, pickle_file)

        coords = [RFT[i, :2] for i in range(RFT.shape[0])]
        all_angles = np.array(flatten([sig for sig, _, _, _, _, _, _ in results]))
        hist, bin_edges = np.histogram(all_angles[:, 0], range=(0, 180), weights=all_angles[:, 1])

        fig, ax1 = plt.subplots()
        ax2 = ax1.twinx()
        ax1.plot(angles, all_data)
        ax1.scatter(angles[(main_peaks + offset) % len(angles)], rolled_all[main_peaks], c="r", label="peaks")
        ax2.stairs(hist, bin_edges)
        ax1.set_ylabel("Intensity Data")
        ax2.set_ylabel("Intensity Binned Angles")
        ax1.set_xlabel("angle (°)")
        plt.tight_layout()
        plt.savefig(os.path.join(results_path, f"{filename}_all_data.{args.image_format}"))
        if args.show_graph: fig.show()

        main_peaks = ((main_peaks + offset) / p_per_a) % 180
        main_peaks = main_peaks.tolist()

        xs, ys = [np.unique(l) for l in zip(*coords)]
        ratio = len(xs) / len(ys)
        grid = [[None for _ in range(len(xs))] for _ in range(len(ys))]
        matched_angles_map = [[[[] for _ in range(len(xs))] for _ in range(len(ys))] for _ in range(len(main_peaks))]
        sig_angles_map = [[[] for _ in range(len(xs))] for _ in range(len(ys))]
        all_angles_map = [[[] for _ in range(len(xs))] for _ in range(len(ys))]
        all_matched_angles_map = [[[] for _ in range(len(xs))] for _ in range(len(ys))]
        img_intensity_map = [[[] for _ in range(len(xs))] for _ in range(len(ys))]
        min_val_map = [[[] for _ in range(len(xs))] for _ in range(len(ys))]
        peak_angle_map = [[[] for _ in range(len(xs))] for _ in range(len(ys))]
        peak_intensity_map = [[[] for _ in range(len(xs))] for _ in range(len(ys))]

        for result in results:
            res = list(result)
            idx = int(res.pop())
            x = find_first(xs, RFT[idx, 0])
            y = find_first(ys, RFT[idx, 1])

            # select: sig pars, bg pars, std error, residuals, min_val, image median intensity, main peaks match

            res += [match(res[0], main_peaks, args.match_angle)]
            grid[y][x] = res

            img_intensity_map[y][x] = res[5]
            min_val_map[y][x] = res[4]
            all_angles_map[y][x] = [lst for lst in res[0] + res[1] if len(lst) > 0]  # All peaks
            sig_angles_map[y][x] = res[0]  # All significant (and unmatched) peaks
            peak_angle_map[y][x] = res[0][np.array(res[0])[:, 1].argmax()][0] if res[0] else np.nan
            peak_intensity_map[y][x] = np.sum([lst[1] for lst in res[0] + res[1] if len(lst) > 0])  # All peaks
            if not np.isnan(res[3]):
                for j, idx in enumerate(res[6]):
                    if not np.isnan(idx):
                        idx = int(idx)
                        a, intensity, s = res[0][j]

                        p = main_peaks[idx]
                        if not abs(a - p) < args.match_angle and abs(
                                (a - p) % 180) < args.match_angle:  # we have a 0 and 180 or vice versa vase
                            a += 180 if a < p else -180
                        if not matched_angles_map[idx][y][x]: # initiate / append to list
                            matched_angles_map[idx][y][x] = [[a, intensity, s]]
                        else:
                            matched_angles_map[idx][y][x] += [[a, intensity, s]]

                        if not all_matched_angles_map[y][x]:
                            all_matched_angles_map[y][x] = [[a, intensity, s]]
                        else:
                            all_matched_angles_map[y][x] += [[a, intensity, s]]

        skip_main_peaks = []  # which main peaks contain 0 data
        for i in range(len(main_peaks)):
            if not any(flatten(matched_angles_map[i])):
                skip_main_peaks.append(i)

        with warnings.catch_warnings(action="ignore"):  # supressing annoying warnings about empty slices
            angles_avg_map = [
                [[angular_weighted_nanmean([d[0] for d in matched_angles_map[i][y][x]]) for x in range(len(xs))] for y
                 in
                 range(len(ys))] for
                i in
                range(len(main_peaks))]
            int_avg_map = [
                [[np.nanmean([d[1] for d in matched_angles_map[i][y][x]]) for x in range(len(xs))] for y in
                 range(len(ys))] for
                i in
                range(len(main_peaks))]
            std_avg_map = [
                [[weighted_nanmean([d[2] for d in matched_angles_map[i][y][x]],
                                   [d[1] for d in matched_angles_map[i][y][x]]) for x in range(len(xs))] for y in
                 range(len(ys))] for
                i in
                range(len(main_peaks))]

        int_map_max = np.nanmax(int_avg_map)

        cmap_viridis = mpl.colormaps.get_cmap('viridis')  # viridis is the default colormap for imshow
        cmap_viridis.set_bad(color='black')
        cmap_hsv = mpl.colormaps.get_cmap('hsv')  # viridis is the default colormap for angle maps
        cmap_hsv.set_bad(color='black')

        row_offset = 0
        lines = [["angle", "mean ang", "std ang", "mean int", "std int", "mean std", "std std", "characteristic_size"]]
        print(lines[-1])
        stats_file.write(", ".join(lines[-1]) + "\n")
        characteristic_sizes = []
        characteristic_sizes_weights = []
        for i, ang_map in enumerate(angles_avg_map):
            if i in skip_main_peaks:
                continue
            fig, (ax1, ax2) = plt.subplots(1, 2, dpi=200, figsize=(6.4 * ratio, 4.8 * 0.65))
            heatmap(np.array(ang_map), ax_imshow=ax1, cbarlabel="angle (°)", cbar_yticks=[0, 45, 90, 135, 180], vmin=0,
                    vmax=180, cmap=cmap_hsv)
            heatmap(np.array(int_avg_map[i]), ax_imshow=ax2, cbarlabel="Intensity", cmap=cmap_viridis, vmin=0,
                    vmax=int_map_max * 0.8)
            fig.tight_layout()
            plt.title(f"Angle {main_peaks[i]:.2f}")
            plt.savefig(os.path.join(results_path, f"{filename}_{main_peaks[i]:.2f}_angle.{args.image_format}"))
            angular_std = np.nanmean(std_avg_map[i])
            characteristic_sizes.append(characteristic_size(angular_std, window_size))
            characteristic_sizes_weights.append(np.nanmean(int_avg_map[i]))
            lines.append([f"{main_peaks[i]}", f"{angular_weighted_nanmean(ang_map, int_avg_map[i])}",
                          f"{angular_std}",
                          f"{np.nanmean(int_avg_map[i])}", f"{np.nanstd(int_avg_map[i])}",
                          f"{np.nanmean(std_avg_map[i])}", f"{np.nanstd(std_avg_map[i])}",
                          f"{characteristic_sizes[-1]}"
            ])
            print(lines[-1])
            stats_file.write(", ".join(lines[-1]) + "\n")
            np.savetxt(os.path.join(results_path, "raw", f"{filename}_{main_peaks[i]:.0f}_angle_raw.csv"),
                       np.array(ang_map), delimiter=",")
            np.savetxt(os.path.join(results_path, "raw", f"{filename}_{main_peaks[i]:.0f}_std_raw.csv"),
                       np.array(std_avg_map[i]), delimiter=",")
            np.savetxt(os.path.join(results_path, "raw", f"{filename}_{main_peaks[i]:.0f}_intensity_raw.csv"),
                       np.array(int_avg_map[i]), delimiter=",")
            if args.show_graph:
                plt.show(block=False)
                plt.pause(0.001)

        stats_file.write("\n")
        lines.append(["characteristic size (px)", f"{weighted_nanmean(characteristic_sizes, characteristic_sizes_weights)}"])
        print(lines[-1])
        stats_file.write(", ".join(lines[-1]) + "\n\n")

        for row in range(1, len(lines) + 1):
            for column in range(1, len(lines[row-1]) + 1):
                try:
                    data = float(lines[row - 1][column - 1])
                except ValueError:
                    data = lines[row - 1][column - 1]
                excel_sheet.cell(row=row + row_offset, column=column, value=data)
        row_offset += 1 + len(lines)

        fig, (ax1, ax2, ax3, ax4) = plt.subplots(1, 4, dpi=200, figsize=(6.4 * ratio * 2, 4.8 * 0.65))
        heatmap(np.array(img_intensity_map), ax_imshow=ax1, cbarlabel="Intensity", cmap=cmap_viridis, vmin=0)
        heatmap(np.array(min_val_map), ax_imshow=ax2, cbarlabel="Min Val", cmap=cmap_viridis, vmin=0)
        heatmap(np.array(peak_angle_map), ax_imshow=ax3, cbarlabel="Peak Angle", cbar_yticks=[0, 45, 90, 135, 180],
                cmap=cmap_hsv, vmin=0, vmax=180)
        heatmap(np.array(peak_intensity_map), ax_imshow=ax4, cbarlabel="Sum Intensities Peak", vmin=0,
                cmap=cmap_viridis)
        fig.tight_layout()
        ax1.set_title(f"Image Median Intensity")
        ax2.set_title(f"Min Val")
        ax3.set_title(f"Peak Angle")
        ax3.set_title(f"Sum Intensities")
        plt.savefig(os.path.join(results_path, f"{filename}_overall.{args.image_format}"))
        if args.show_graph:
            plt.show(block=False)
            plt.pause(0.001)
        np.savetxt(os.path.join(results_path, "raw", f"{filename}_peak_angle_raw.csv"), np.array(peak_angle_map),
                   delimiter=",")
        np.savetxt(os.path.join(results_path, "raw", f"{filename}_peak_intensity_raw.csv"),
                   np.array(peak_intensity_map), delimiter=",")

        # Parameters
        neigh_sizes = np.arange(3, min([len(xs), len(ys), args.max_neighbourhood + 2]), 2, dtype=int)

        labels = ["ang order", "ang std order", "std intensity", "intensity", "intensity+min_val",
                  "int*min_val*ang_order"]
        labels_all = ["all angles", "sig angles", "peak angles"]
        vmins = [0, 0, 0, 0, 0, 0]
        vmaxs = [1, 1, None, None, None, None]

        if args.all_angles:
            todo = list(main_peaks) + ["all"]
        else:
            todo = ["all"]
            skip_main_peaks = []

        for i, ang in enumerate(todo):
            if i in skip_main_peaks:
                continue
            if neigh_sizes.size == 0:
                continue

            if type(ang) != str:
                ang = f"{ang:.2f}"

            stats_file = open(os.path.join(results_path, "csv", f"{filename}_{ang}_stats.csv"), "a")

            fig, axes = plt.subplots(len(neigh_sizes) if len(neigh_sizes) > 1 else 2,
                                     len(labels) if ang != "all" else len(labels_all), dpi=400,
                                     figsize=(
                                         6.4 * 2.5 * (1.4 if ang != "all" else 0.7), 4.8 * 2 * len(neigh_sizes) / 6))
            fig.text(0.015, 0.5, f"Angle\n{ang}", fontsize=14)
            text = [f"{ang}: Neighbourhood size "] + flatten(
                [["mean " + lbl, "std " + lbl] for lbl in (labels if ang != "all" else labels_all)])
            stats_file.write(", ".join(text) + "\n")
            lines = [text]
            for n_index in range(len(neigh_sizes)):
                neigh_size = neigh_sizes[n_index]
                print(f"{i}, {ang}, Neigh size: {neigh_size}")
                stats_file.write(f"{neigh_size}")
                line = [f"{neigh_size}"]
                if ang == "all":
                    order_params = [[[] for _ in range(len(xs))] for _ in range(len(ys))]
                    for map_ref in [all_angles_map, sig_angles_map, all_matched_angles_map]:
                        for y in range(args.filter_edges, len(ys) - neigh_size + 1 - args.filter_edges):
                            for x in range(args.filter_edges, len(xs) - neigh_size + 1 - args.filter_edges):
                                all_angles_local = [map_ref[y + dy][x:x + neigh_size] for dy in range(neigh_size)]
                                centre = all_angles_local[neigh_size // 2].pop(neigh_size // 2)

                                if bool(centre and flatten(flatten(all_angles_local))):
                                    angs, ints, _ = [np.array(lst) for lst in zip(*flatten(flatten(all_angles_local)))]
                                    order_params[y + neigh_size // 2][x + neigh_size // 2].append(
                                        calc_order_param_weighted(angs, ints, centre))
                                else:
                                    order_params[y + neigh_size // 2][x + neigh_size // 2].append(np.nan)

                    for imshow in range(len(labels_all)):
                        curr_map = np.array([[v[imshow] if v else np.nan for v in vs] for vs in order_params])
                        im, cbar = heatmap(curr_map, ax_imshow=axes[n_index, imshow], vmin=0, vmax=1, cmap=cmap_viridis)
                        mean = np.nanmean(curr_map)
                        std = np.nanstd(curr_map)
                        print(f"{labels_all[imshow]}: mean: {mean:.2f}, std: {std:.2f}")
                        stats_file.write(f", {mean:.2f}, {std:.2f}")
                        line += [f"{mean:.2f}", f"{std:.2f}"]
                        np.savetxt(os.path.join(results_path, "raw",
                                                f"{filename}_order_{labels_all[imshow]}_neighbourhood_{neigh_sizes[n_index]}_raw.csv"),
                                   np.array(curr_map), delimiter=",")

                    for ax, col in zip(axes[0], labels_all):
                        ax.set_title(col)
                else:

                    order_params = [[np.full(len(labels), np.nan) for _ in range(len(xs))] for _ in range(len(ys))]
                    for y in range(args.filter_edges, len(ys) - neigh_size + 1 - args.filter_edges):
                        for x in range(args.filter_edges, len(xs) - neigh_size + 1 - args.filter_edges):
                            # select: sig pars, bg pars, std error, residuals, min_val
                            all_angles_local = [matched_angles_map[i][y + dy][x:x + neigh_size] for dy in
                                                range(neigh_size)]
                            min_vals = [min_val_map[y + dy][x:x + neigh_size] for dy in
                                        range(neigh_size)]

                            centre_value = all_angles_local[neigh_size // 2].pop(neigh_size // 2)
                            all_angles_local = flatten(flatten(all_angles_local))
                            if bool(centre_value and all_angles_local):
                                angs, intensities, standards = [np.array(d) for d in zip(*all_angles_local)]

                                # centre cell x, y = angle ord, std of angle ord, intensity of angle ord, base int ord
                                order_pars = calc_order_param(angs, intensities, standards, centre_value)
                                order_params[y + neigh_size // 2][x + neigh_size // 2] = order_pars + [
                                    np.nanmean(intensities),
                                    np.nanmean(intensities) + np.mean(min_vals),
                                    np.nanmean(intensities) * np.mean(min_vals) * order_pars[0]
                                ]

                    # print(order_params)
                    if vmaxs[2] is None:
                        for idx in range(len(vmaxs)):
                            if vmaxs[idx] is None:
                                vmaxs[idx] = np.nanmax(np.array([[v[idx] for v in vs] for vs in order_params]))

                    for imshow in range(len(labels)):
                        curr_map = np.array([[v[imshow] for v in vs] for vs in order_params])
                        im, cbar = heatmap(curr_map, ax_imshow=axes[n_index, imshow], vmin=vmins[imshow],
                                           vmax=vmaxs[imshow],
                                           cmap=cmap_viridis)
                        mean = weighted_nanmean(curr_map)
                        std = weighted_nanstd(curr_map)
                        print(f"{labels[imshow]}: mean: {mean:.2f}, std: {std:.2f}")
                        stats_file.write(f", {mean:.2f}, {std:.2f}")
                        line += [f"{mean:.2f}", f"{std:.2f}"]
                        np.savetxt(os.path.join(results_path, "raw", f"{filename}_order_{labels[imshow]}_raw.csv"),
                                   np.array(curr_map), delimiter=",")

                    for ax, col in zip(axes[0], labels):
                        ax.set_title(col)

                    for ax, row in zip(axes[:, 0], neigh_sizes):
                        ax.set_ylabel(row, rotation=0, size='large')

                lines.append(line)
                stats_file.write("\n")


            plt.tight_layout()
            plt.savefig(os.path.join(results_path, f"{filename}_order_{ang}.{args.image_format}"))
            if args.show_graph:
                plt.show(block=False)
                plt.pause(0.001)

            excel_sheet.append(["", ""])
            for row in range(1, len(lines) + 1):
                for column in range(1, len(lines[0]) + 1):
                    try:
                        data = float(lines[row - 1][column - 1])
                    except ValueError:
                        data = lines[row - 1][column - 1]
                    excel_sheet.cell(row=row + row_offset, column=column, value=data)
            row_offset += 1 + len(lines)
            stats_file.close()

        # Profile over flow axis
        vert_profile = np.nansum(peak_intensity_map, 1)
        hori_profile = np.nansum(peak_intensity_map, 0)

        fig, (ax1, ax2) = plt.subplots(nrows=1, ncols=2)
        ax1.plot(vert_profile)
        ax2.plot(hori_profile)
        ax1.set_title("Vertical Profile")
        ax2.set_title("Horizontal Profile")

        if args.show_graph: fig.show()
        fig.savefig(os.path.join(results_path, f"{filename}_flow_profile.{args.image_format}"))

        excel_sheet.cell(row=row_offset, column=1, value="Vertical Profile")
        for i, val in enumerate(vert_profile):
            excel_sheet.cell(row=row_offset + 1, column=i + 1, value=val)

        excel_sheet.cell(row=row_offset + 2, column=1, value="Horizontal Profile")
        for i, val in enumerate(hori_profile):
            excel_sheet.cell(row=row_offset + 3, column=i + 1, value=val)

        if not args.show_graph: plt.close('all')
        np.savetxt(os.path.join(results_path, "raw", f"{filename}_intensity_map_raw.csv"),
                   np.array(img_intensity_map), delimiter=",")

    if args.absolute:
        excel_path = os.path.join(args.core_path, "results", *folder_path.split(os.sep)[args.core_path.count(os.sep) + 2:],
                                  f"{os.path.basename(args.core_path) if args.core_path and len(filepaths)>1 else filename}_overview.xlsx")
    else:
        excel_path =os.path.join(".", "results", args.core_path if args.core_path else filename,
                                     f"{os.path.basename(args.core_path) if args.core_path and len(filepaths)>1 else filename}_overview.xlsx")
    excel_book.save(excel_path)

    print("Saving results to", excel_path)
    if not args.IDE and args.show_graph: plt.show()
if __name__ == "__main__":
    run()